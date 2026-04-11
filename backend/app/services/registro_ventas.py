"""
Registro de Ventas — genera un Excel mensual para enviar a la contadora.

Formato estándar de Registro de Ventas e Ingresos para SUNAT PLE:
- Encabezado: REGISTRO DE VENTAS, RUC, Razón Social, Usuario, AL: última fecha
- Columnas: N°, F.Emisión, F.Vcto, TC, N.Serie, N.Doc, TD, Número, Cliente,
  Op.Gravada, Op.Inafecta, Op.Exonerada, IGV, Otros Tributos, Total, T/C,
  Fecha_Ref, TC_Ref, Serie_Ref, Numero_Ref, detalle, estado, monto_validacion.
"""
from __future__ import annotations

import io
from calendar import monthrange
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload

from app.config import settings
from app.models.sale import Sale

# SUNAT document type codes (Catálogo 10)
DOC_TYPE_SUNAT = {
    "FACTURA": "01",
    "BOLETA": "03",
    "NOTA_CREDITO": "07",
    "NOTA_DEBITO": "08",
    "NOTA_VENTA": "00",  # Internal, not fiscal
}

# SUNAT client doc type codes (Catálogo 06)
CLIENT_DOC_SUNAT = {
    "RUC": "6",
    "DNI": "1",
    "NONE": "0",
    "CE": "4",
    "PASAPORTE": "7",
}

MONTH_NAMES_ES = [
    "", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE",
]


def get_monthly_sales(db: Session, year: int, month: int) -> list[Sale]:
    """Retrieve all FACTURADO and ANULADO sales for the given month (Lima time)."""
    last_day = monthrange(year, month)[1]
    start = date(year, month, 1)
    end = date(year, month, last_day)

    lima_date = func.date(func.timezone("America/Lima", Sale.created_at))

    return (
        db.query(Sale)
        .options(joinedload(Sale.client), joinedload(Sale.ref_sale))
        .filter(
            Sale.status.in_(["FACTURADO", "ANULADO"]),
            Sale.doc_type.in_(["FACTURA", "BOLETA", "NOTA_CREDITO"]),
            lima_date >= start,
            lima_date <= end,
        )
        .order_by(Sale.doc_type, Sale.series, Sale.doc_number)
        .all()
    )


def generate_registro_ventas_xlsx(db: Session, year: int, month: int) -> bytes:
    """Generate the Registro de Ventas XLSX for the given month."""
    sales = get_monthly_sales(db, year, month)
    last_day = monthrange(year, month)[1]
    al_date = f"{last_day:02d}/{MONTH_NAMES_ES[month]}/{year}"

    wb = Workbook()
    ws = wb.active
    ws.title = f"Registro {year}{month:02d}"

    # ─── Header block (rows 1-5) ───
    ws["A1"] = "REGISTRO DE VENTAS"
    ws["A1"].font = Font(bold=True, size=12)

    ws["A2"] = f"RUC: {settings.EMPRESA_RUC}"
    ws["A2"].font = Font(bold=True)

    ws["A3"] = settings.EMPRESA_RAZON_SOCIAL
    ws["A3"].font = Font(bold=True)

    ws["A4"] = "USUARIO: KTI"
    ws["A4"].font = Font(bold=True)

    ws["A5"] = f"AL: {al_date}"
    ws["A5"].font = Font(bold=True)

    # ─── Column headers (row 8) ───
    headers = [
        "N°",              # 1
        "F. Emisión",      # 2
        "F. Vcto.",        # 3
        "TC.",             # 4  tipo comprobante
        "N. Serie",        # 5
        "N. Doc",          # 6
        "TD.",             # 7  tipo doc cliente
        "Número",          # 8  número doc cliente
        "Cliente",         # 9
        "Op. Grav",        # 10 operación gravada
        "Op. Inaf",        # 11 operación inafecta
        "Op. Exo",         # 12 operación exonerada
        "IGV",             # 13
        "Otros Trib.",     # 14 otros tributos
        "Total",           # 15
        "T/C",             # 16 tipo de cambio
        "Fecha_Ref",       # 17 fecha referencia (NC)
        "TC_Ref",          # 18 tipo comprobante referencia (NC)
        "Serie_Ref",       # 19 serie referencia (NC)
        "Numero_Ref",      # 20 numero referencia (NC)
        "detalle",         # 21
        "estado",          # 22
        "monto_validacion",# 23
    ]

    header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    HEADER_ROW = 8
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # ─── Data rows ───
    total_gravada = 0.0
    total_inafecta = 0.0
    total_exonerada = 0.0
    total_igv = 0.0
    total_otros = 0.0
    total_importe = 0.0

    row = HEADER_ROW + 1
    for idx, sale in enumerate(sales, start=1):
        issue_date = sale.issue_date or (
            sale.created_at.date() if sale.created_at else None
        )

        doc_tipo_sunat = DOC_TYPE_SUNAT.get(sale.doc_type, "")
        client = sale.client
        client_doc_type_raw = (client.doc_type if client else "") or ""
        client_doc_type = CLIENT_DOC_SUNAT.get(client_doc_type_raw, "0")
        client_doc_num = (client.doc_number if client else "") or ""
        client_name = (client.business_name if client else "") or ""

        # For boletas to walk-in customers (no RUC), use "CLIENTES VARIOS"
        if sale.doc_type == "BOLETA" and client_doc_type_raw != "RUC":
            client_name = "CLIENTES VARIOS"
            client_doc_num = "00000000"
            client_doc_type = "0"

        # NC stores amounts as positive in the DB — show negative in the registry
        is_nc = sale.doc_type == "NOTA_CREDITO"
        sign = -1 if is_nc else 1

        base = float(sale.subtotal) * sign
        igv = float(sale.igv_amount) * sign
        total = float(sale.total) * sign

        # Estado: 1 = Anotado, 9 = Anulado (SUNAT PLE codes)
        estado_code = 9 if sale.status == "ANULADO" else 1

        # NC reference fields
        ref_fecha = ""
        ref_tc = ""
        ref_serie = ""
        ref_num = ""
        if is_nc and sale.ref_sale:
            ref_sale = sale.ref_sale
            if ref_sale.issue_date:
                ref_fecha = ref_sale.issue_date.strftime("%d/%m/%Y")
            elif ref_sale.created_at:
                ref_fecha = ref_sale.created_at.date().strftime("%d/%m/%Y")
            ref_tc = DOC_TYPE_SUNAT.get(ref_sale.doc_type, "")
            ref_serie = ref_sale.series or ""
            ref_num = str(ref_sale.doc_number).zfill(8) if ref_sale.doc_number else ""

        doc_num_str = str(sale.doc_number).zfill(8) if sale.doc_number else "-"

        values = [
            idx,                                                       # 1 N°
            issue_date.strftime("%d/%m/%Y") if issue_date else "",     # 2 F. Emisión
            "",                                                         # 3 F. Vcto
            doc_tipo_sunat,                                             # 4 TC
            sale.series,                                                # 5 N. Serie
            doc_num_str,                                                # 6 N. Doc
            client_doc_type,                                            # 7 TD
            client_doc_num,                                             # 8 Número
            client_name,                                                # 9 Cliente
            round(base, 2),                                             # 10 Op. Grav
            0.00,                                                       # 11 Op. Inaf
            0.00,                                                       # 12 Op. Exo
            round(igv, 2),                                              # 13 IGV
            0.00,                                                       # 14 Otros Trib.
            round(total, 2),                                            # 15 Total
            1.000,                                                      # 16 T/C (PEN)
            ref_fecha,                                                  # 17 Fecha_Ref
            ref_tc,                                                     # 18 TC_Ref
            ref_serie,                                                  # 19 Serie_Ref
            ref_num,                                                    # 20 Numero_Ref
            "",                                                         # 21 detalle
            estado_code,                                                # 22 estado
            "",                                                         # 23 monto_validacion
        ]

        NUMERIC_COLS = {10, 11, 12, 13, 14, 15, 16}
        CENTER_COLS = {1, 4, 5, 6, 7, 8, 18, 19, 22}

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = border
            if col_idx == 16:
                cell.number_format = '0.000'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in NUMERIC_COLS:
                cell.number_format = '#,##0.00;[Red]-#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in CENTER_COLS:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 9:
                cell.alignment = Alignment(horizontal="left")

        # Accumulate totals only for non-voided rows
        if sale.status != "ANULADO":
            total_gravada += base
            total_igv += igv
            total_importe += total

        row += 1

    # ─── Totals row ───
    if row > HEADER_ROW + 1:
        total_row = row + 1
        tot_label = ws.cell(row=total_row, column=9, value="TOTALES")
        tot_label.font = Font(bold=True)
        tot_label.alignment = Alignment(horizontal="right")
        tot_label.fill = PatternFill(start_color="E8ECF2", end_color="E8ECF2", fill_type="solid")

        for col_idx, val in (
            (10, total_gravada),
            (11, total_inafecta),
            (12, total_exonerada),
            (13, total_igv),
            (14, total_otros),
            (15, total_importe),
        ):
            c = ws.cell(row=total_row, column=col_idx, value=round(val, 2))
            c.font = Font(bold=True)
            c.number_format = '#,##0.00;[Red]-#,##0.00'
            c.alignment = Alignment(horizontal="right")
            c.fill = PatternFill(start_color="E8ECF2", end_color="E8ECF2", fill_type="solid")
            c.border = border

    # ─── Column widths ───
    widths = [
        5,    # N°
        12,   # F. Emisión
        12,   # F. Vcto
        6,    # TC
        10,   # N. Serie
        12,   # N. Doc
        6,    # TD
        14,   # Número cliente
        40,   # Cliente
        12,   # Op. Grav
        10,   # Op. Inaf
        10,   # Op. Exo
        11,   # IGV
        11,   # Otros Trib.
        12,   # Total
        8,    # T/C
        12,   # Fecha_Ref
        8,    # TC_Ref
        10,   # Serie_Ref
        12,   # Numero_Ref
        14,   # detalle
        8,    # estado
        16,   # monto_validacion
    ]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[HEADER_ROW].height = 28
    ws.freeze_panes = f"A{HEADER_ROW + 1}"

    # ─── Save to bytes ───
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def build_filename(year: int, month: int) -> str:
    """Standard filename for the monthly registry."""
    return f"Registro_Ventas_{settings.EMPRESA_RUC}_{year}{month:02d}.xlsx"
